"""Geração do arquivo Excel (.xlsx) final do relatório financeiro, com abas
formatadas (cabeçalho, moeda, autofiltro, congelamento de painel)."""
from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MOEDA_FMT = "R$ #,##0.00"

_COLUNAS_MOEDA_PADRAO = {
    "Valor Título", "Valor Pago/Recebido", "Valor Aberto", "Juros", "Multa",
    "Desconto", "Valor Líquido", "Pagar (Aberto)", "Receber (Aberto)", "Saldo",
}

# Estilo da aba "Geral", padronizado com o modelo de exportação nativo da Omie
# (aba "bdContas"): cabeçalho cinza claro com quebra de linha, valores em formato
# contábil (parênteses para negativo) e datas em dd/mm/yyyy.
_GERAL_HEADER_FILL = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
_GERAL_HEADER_FONT = Font(color="111111", bold=True)
_CONTABIL_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
_DATA_FMT = "dd/mm/yyyy"

_COLUNAS_GERAL_MOEDA = {
    "Valor da Conta", "Pago ou Recebido", "A Pagar ou Receber",
    "COFINS Retido", "CSLL Retido", "INSS Retido", "IR Retido", "ISS Retido", "PIS Retido",
    "Desconto", "Juros",
}
_COLUNAS_GERAL_DATA = {
    "Data de Registro (completa)", "Data de Emissão (completa)",
    "Data de Vencimento (completa)", "Data de Pagto",
}


def _formatar_planilha(ws: Worksheet, df: pd.DataFrame) -> None:
    if df.empty:
        return

    for col_idx, nome_col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    is_moeda = [nome_col in _COLUNAS_MOEDA_PADRAO for nome_col in df.columns]
    for row_idx in range(2, len(df) + 2):
        for col_idx, moeda in enumerate(is_moeda, start=1):
            if moeda:
                ws.cell(row=row_idx, column=col_idx).number_format = _MOEDA_FMT

    for col_idx, nome_col in enumerate(df.columns, start=1):
        try:
            maior_valor = df.iloc[:, col_idx - 1].astype(str).map(len).max()
        except (ValueError, TypeError):
            maior_valor = 10
        largura = min(max(len(str(nome_col)), int(maior_valor or 0)) + 3, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"


def _formatar_geral(ws: Worksheet, df: pd.DataFrame) -> None:
    """Formata a aba "Geral" no estilo do modelo padrão de exportação da Omie:
    cabeçalho cinza claro em negrito com quebra de linha, valores em formato
    contábil e datas em dd/mm/yyyy."""
    if df.empty:
        return

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _GERAL_HEADER_FILL
        cell.font = _GERAL_HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    is_moeda = [nome_col in _COLUNAS_GERAL_MOEDA for nome_col in df.columns]
    is_data = [nome_col in _COLUNAS_GERAL_DATA for nome_col in df.columns]
    for row_idx in range(2, len(df) + 2):
        for col_idx, (moeda, data_col) in enumerate(zip(is_moeda, is_data), start=1):
            if moeda:
                ws.cell(row=row_idx, column=col_idx).number_format = _CONTABIL_FMT
            elif data_col:
                ws.cell(row=row_idx, column=col_idx).number_format = _DATA_FMT

    for col_idx, nome_col in enumerate(df.columns, start=1):
        try:
            maior_valor = df.iloc[:, col_idx - 1].astype(str).map(len).max()
        except (ValueError, TypeError):
            maior_valor = 10
        largura = min(max(len(str(nome_col)), int(maior_valor or 0)) + 3, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"


def _escrever_resumo(ws: Worksheet, resumo: dict[str, Any]) -> None:
    ws.append(["Indicador", "Valor"])
    for col_idx in (1, 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    campos_moeda = {k for k in resumo if "Qtd" not in k}
    for label, valor in resumo.items():
        ws.append([label, valor])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True)
        if label in campos_moeda:
            ws.cell(row=row, column=2).number_format = _MOEDA_FMT

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A2"


def escrever_planilha_movimentacoes(
    caminho_saida: str, df_geral: pd.DataFrame, nome_aba: str = "Movimentações"
) -> None:
    """Grava só a aba de movimentações (mesmo layout/formatação de `_formatar_geral`,
    usado pela aba "Geral" dos relatórios completos), sem as abas analíticas
    (Resumo/Por Status/Por Categoria/etc.) — usado por `main_planilha.py` para
    manter uma planilha única e estável com o snapshot mais atual das
    movimentações, sobrescrita por completo a cada execução."""
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_geral.to_excel(writer, sheet_name=nome_aba, index=False)
        _formatar_geral(writer.book[nome_aba], df_geral)


def escrever_relatorio(
    caminho_saida: str,
    resumo: dict[str, Any],
    df_geral: pd.DataFrame,
    df_pagar: pd.DataFrame,
    df_receber: pd.DataFrame,
    df_status: pd.DataFrame,
    df_categoria: pd.DataFrame,
    df_cliente: pd.DataFrame,
    df_fluxo: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        # Aba "Resumo" é escrita manualmente (chave/valor), as demais via pandas.
        pd.DataFrame().to_excel(writer, sheet_name="Resumo", index=False)
        df_geral.to_excel(writer, sheet_name="Geral", index=False)
        df_pagar.to_excel(writer, sheet_name="Contas a Pagar", index=False)
        df_receber.to_excel(writer, sheet_name="Contas a Receber", index=False)
        df_status.to_excel(writer, sheet_name="Por Status", index=False)
        df_categoria.to_excel(writer, sheet_name="Por Categoria", index=False)
        df_cliente.to_excel(writer, sheet_name="Por Cliente-Fornecedor", index=False)
        df_fluxo.to_excel(writer, sheet_name="Fluxo Mensal", index=False)

        wb = writer.book
        _escrever_resumo(wb["Resumo"], resumo)
        _formatar_geral(wb["Geral"], df_geral)
        _formatar_planilha(wb["Contas a Pagar"], df_pagar)
        _formatar_planilha(wb["Contas a Receber"], df_receber)
        _formatar_planilha(wb["Por Status"], df_status)
        _formatar_planilha(wb["Por Categoria"], df_categoria)
        _formatar_planilha(wb["Por Cliente-Fornecedor"], df_cliente)
        _formatar_planilha(wb["Fluxo Mensal"], df_fluxo)
